import requests

# Input Excel file and output log file
input_file = "URL_testing.xlsx"
log_file = "url_status.log"

# Function to read URLs from the Excel file
def read_urls_from_excel(file_name):
    import openpyxl
    wb = openpyxl.load_workbook(file_name)
    if "Sheet1" not in wb.sheetnames:
        raise ValueError("Sheet1 is missing in the workbook.")
    sheet1 = wb["Sheet1"]
    urls = [row[0] for row in sheet1.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
    return urls

# Function to check if URLs are working and log results
def check_urls(urls, log_file):
    total_urls = len(urls)
    with open(log_file, "w") as log:
        log.write("URL Status Check Results\n")
        log.write("-" * 50 + "\n")
        log.write(f"{'URL':<50} {'Status'}\n")
        log.write("-" * 50 + "\n")

        for i, url in enumerate(urls, start=1):
            try:
                response = requests.get(url, timeout=10)
                if response.status_code == 200:
                    status = "Success"
                else:
                    status = f"Fail (HTTP {response.status_code})"
            except requests.exceptions.RequestException:
                status = "Fail"

            log.write(f"{url:<50} {status}\n")

            # Print progress update
            print(f"[{i}/{total_urls}] Checked: {url} - {status}")

# Main execution
if __name__ == "__main__":
    try:
        urls = read_urls_from_excel(input_file)
        print(f"Found {len(urls)} URLs to check...")
        check_urls(urls, log_file)
        print(f"URL status check completed. Results saved to {log_file}")
    except Exception as e:
        print(f"Error: {e}")
